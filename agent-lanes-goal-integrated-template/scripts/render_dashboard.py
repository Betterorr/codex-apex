#!/usr/bin/env python3
"""Render a human-readable Agent Lanes dashboard and communication workbook."""

from __future__ import annotations

import json
import re
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any


ROOT = Path(__file__).resolve().parents[2]
AGENT_LANES = ROOT / "agent-lanes"

LANE_LABELS = {
    "orchestrator": "主调度泳道",
    "planning": "规划泳道",
    "design": "设计泳道",
    "development": "开发泳道",
    "guardian": "守门泳道",
    "review": "验收泳道",
    "evolution": "进化泳道",
    "callback_post_office": "回报邮局",
}

STATUS_LABELS = {
    "active": "运行中",
    "paused": "暂停",
    "archived": "归档",
    "blocked": "阻塞",
}

TRANSPORT_TASK_TYPES = {"callback_batch_ready", "callback_spooled", "orchestrator_state"}


def now_text() -> str:
    return datetime.now(timezone(timedelta(hours=8))).replace(microsecond=0).isoformat()


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8-sig"))


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if not path.exists():
        return rows
    for line_no, line in enumerate(path.read_text(encoding="utf-8-sig", errors="replace").splitlines(), start=1):
        text = line.strip()
        if not text:
            continue
        try:
            rows.append(json.loads(text))
        except json.JSONDecodeError:
            rows.append(
                {
                    "message_id": f"parse-error-line-{line_no}",
                    "status": "PARSE_ERROR",
                    "summary": text[:160],
                    "created_at": "",
                }
            )
    return rows


def rel(path: str | Path) -> str:
    path_obj = Path(path)
    text = path_obj.as_posix()
    if path_obj.is_absolute():
        try:
            return path_obj.relative_to(ROOT).as_posix()
        except ValueError:
            return text
    return text


def md_link(label: str, path: str | Path) -> str:
    return f"[{label}]({rel(path)})"


def readable(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def compact(value: Any, max_len: int = 120) -> str:
    text = readable(value).replace("\r", " ").replace("\n", " ")
    text = re.sub(r"\s+", " ", text).strip()
    if len(text) <= max_len:
        return text
    return text[: max_len - 1].rstrip() + "..."


def csv_value(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, list):
        return "\n".join(csv_value(item) for item in value)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    return str(value)


def time_text(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    normalized = text.replace("Z", "+00:00")
    try:
        parsed = datetime.fromisoformat(normalized)
    except ValueError:
        return text
    return parsed.strftime("%Y-%m-%d %H:%M:%S")


def short_time(value: Any) -> str:
    text = str(value or "")
    if "T" in text:
        return text.split("T", 1)[1][:8]
    return compact(text, 16)


def lane_label(value: Any, aliases: dict[str, str] | None = None) -> str:
    text = str(value or "")
    if aliases and text in aliases:
        return aliases[text]
    return LANE_LABELS.get(text, text or "-")


def status_badge(status: Any) -> str:
    return STATUS_LABELS.get(str(status or ""), str(status or "-"))


def iter_text_values(value: Any) -> list[str]:
    if isinstance(value, dict):
        texts: list[str] = []
        for item in value.values():
            texts.extend(iter_text_values(item))
        return texts
    if isinstance(value, list):
        texts = []
        for item in value:
            texts.extend(iter_text_values(item))
        return texts
    if isinstance(value, str):
        return [value]
    return []


def looks_garbled(value: Any) -> bool:
    mojibake_fragments = [
        "涓",
        "娉抽亾",
        "搴︽",
        "鈥",
        "涔辩爜",
        "鏃堕棿",
        "鍥炴姤",
        "楠屾敹",
        "寮€",
        "宸插",
        "閭",
        "绋",
        "鐨",
        "鏄",
        "鍚",
        "锛",
        "銆",
        "€",
    ]
    for text in iter_text_values(value):
        stripped = text.strip()
        if not stripped:
            continue
        if "???" in stripped or set(stripped) == {"?"}:
            return True
        if "�" in stripped:
            return True
        if any(fragment in stripped for fragment in mojibake_fragments):
            return True
    return False


def message_is_garbled(msg: dict[str, Any]) -> bool:
    fields = [
        msg.get("from_agent"),
        msg.get("to_agent"),
        msg.get("summary"),
        msg.get("evidence"),
        msg.get("concerns"),
        msg.get("next_recommended_action"),
        msg.get("orchestrator_message"),
        msg.get("thread_prompt"),
    ]
    return any(looks_garbled(field) for field in fields)


def completion_like(msg: dict[str, Any]) -> bool:
    if msg.get("task_type") in {"completion_callback", "completion_callback_protocol_update"}:
        return True
    if msg.get("status") not in {"DONE", "DONE_WITH_CONCERNS", "NEEDS_CONTEXT", "BLOCKED"}:
        return False
    fields = ["summary", "changed_files", "evidence", "concerns", "next_recommended_lane", "next_recommended_action"]
    return sum(1 for field in fields if msg.get(field) not in (None, "", [])) >= 4


def valid_display_callback(msg: dict[str, Any]) -> bool:
    return completion_like(msg) and not message_is_garbled(msg)


def is_dispatch(msg: dict[str, Any]) -> bool:
    task_type = str(msg.get("task_type") or "")
    status = str(msg.get("status") or "")
    if status in {"DISPATCHED", "sent", "active_dispatch"}:
        return True
    if status in {"DONE", "DONE_WITH_CONCERNS", "NEEDS_CONTEXT", "BLOCKED"}:
        return False
    return "dispatch" in task_type


def active_slice_dispatch(current_state: dict[str, Any]) -> dict[str, Any] | None:
    active = current_state.get("active_value_slice")
    if not isinstance(active, dict):
        return None
    status = str(active.get("status") or "").upper()
    if status in {"DONE", "DONE_WITH_CONCERNS", "BLOCKED", "CANCELLED", "COMPLETE"}:
        return None
    message_id = active.get("dispatch_message_id")
    target_lane = active.get("target_lane")
    if not message_id or not target_lane:
        return None
    return {
        "message_id": str(message_id),
        "from_agent": "orchestrator",
        "to_agent": str(target_lane),
        "task_type": "active_value_slice_dispatch",
        "status": "DISPATCHED",
        "summary": active.get("title") or active.get("value_slice_id") or "当前 Value Slice",
        "value_slice_id": active.get("value_slice_id"),
        "created_at": active.get("started_at"),
        "source": "agent-lanes/current-state.json",
    }


def is_user_to_orchestrator_request(msg: dict[str, Any], aliases: dict[str, str]) -> bool:
    to_lane = lane_label(msg.get("to_agent"), aliases)
    from_lane = lane_label(msg.get("from_agent"), aliases)
    task_type = str(msg.get("task_type") or "").lower()
    return to_lane == "主调度泳道" and from_lane != "主调度泳道" and (
        from_lane.startswith("用户") or "user" in task_type
    )


def dispatch_reply_tokens(msg: dict[str, Any]) -> set[str]:
    tokens = {str(msg.get("message_id") or "")}
    for key, value in msg.items():
        if key.endswith("_dispatch_id") and value:
            tokens.add(str(value))
    return {token for token in tokens if token}


def is_reply_to(msg: dict[str, Any], dispatch_id: str) -> bool:
    return str(msg.get("reply_to") or "") == dispatch_id and completion_like(msg)


def message_blob(msg: dict[str, Any]) -> str:
    return " ".join(iter_text_values(msg))


def dispatch_is_covered_by_callback(
    messages: list[dict[str, Any]], dispatch: dict[str, Any], lane_name: str, aliases: dict[str, str]
) -> bool:
    tokens = dispatch_reply_tokens(dispatch)
    if not tokens:
        return False
    dispatch_slice_id = str(dispatch.get("value_slice_id") or (dispatch.get("value_slice") or {}).get("value_slice_id") or "")
    for candidate in messages:
        if not completion_like(candidate):
            continue
        if lane_label(candidate.get("from_agent"), aliases) != lane_name:
            continue
        if lane_label(candidate.get("to_agent"), aliases) != "主调度泳道":
            continue
        candidate_slice_id = str(candidate.get("value_slice_id") or "")
        if dispatch_slice_id and candidate_slice_id != dispatch_slice_id:
            continue
        if any(is_reply_to(candidate, token) for token in tokens):
            return True
    return False


def control_plane_conflicts(
    current_state: dict[str, Any], ledger_rows: list[dict[str, Any]], messages: list[dict[str, Any]], aliases: dict[str, str]
) -> list[str]:
    conflicts: list[str] = []
    active = current_state.get("active_value_slice")
    if not isinstance(active, dict):
        return conflicts
    dispatch_id = str(active.get("dispatch_message_id") or "")
    slice_id = str(active.get("value_slice_id") or "")
    target_lane = lane_label(active.get("target_lane"), aliases)
    dispatch_event = next(
        (
            row
            for row in ledger_rows
            if row.get("event_type") == "dispatch_started"
            and dispatch_id in {str(row.get("interaction_id") or ""), str(row.get("source_message_id") or "")}
        ),
        None,
    )
    if not dispatch_event:
        conflicts.append("current-state active dispatch is missing from Value Slice ledger")
        return conflicts
    if str(dispatch_event.get("value_slice_id") or "") != slice_id:
        conflicts.append("current-state value_slice_id conflicts with ledger dispatch")
    if lane_label(dispatch_event.get("target_lane"), aliases) != target_lane:
        conflicts.append("current-state target_lane conflicts with ledger dispatch")
    for candidate in messages:
        if not completion_like(candidate) or str(candidate.get("reply_to") or "") != dispatch_id:
            continue
        if lane_label(candidate.get("from_agent"), aliases) != target_lane:
            conflicts.append("a callback replies to the active dispatch from the wrong lane")
        if str(candidate.get("value_slice_id") or "") != slice_id:
            conflicts.append("a callback replies to the active dispatch with the wrong value_slice_id")
    return sorted(set(conflicts))


def item_link_or_code(item: Any) -> str:
    text = readable(item)
    if not text:
        return "-"
    path = ROOT / text
    if path.exists():
        return md_link(text, text)
    return f"`{text}`" if ("/" in text or "\\" in text or "." in Path(text).name) else text


def append_value_block(lines: list[str], title: str, value: Any) -> None:
    if value in (None, "", []):
        return
    lines.append(f"**{title}**")
    if isinstance(value, list):
        for item in value:
            lines.append(f"- {item_link_or_code(item)}")
    else:
        for paragraph in readable(value).splitlines():
            paragraph = paragraph.strip()
            if paragraph:
                lines.append(paragraph)
    lines.append("")


def communication_type(msg: dict[str, Any]) -> str:
    task_type = str(msg.get("task_type") or "")
    if task_type == "callback_batch_ready":
        return "邮局合并批次"
    if task_type == "callback_spooled":
        return "邮局暂存回执"
    if task_type == "orchestrator_state":
        return "主调度忙闲状态"
    if is_dispatch(msg):
        return "主调度派发"
    if completion_like(msg):
        return "泳道完成回报"
    return "普通记录"


def quality_state(msg: dict[str, Any], aliases: dict[str, str]) -> str:
    if message_is_garbled(msg):
        return "乱码/需重发"
    to_agent = lane_label(msg.get("to_agent"), aliases)
    if completion_like(msg) and not msg.get("delivery_mode") and to_agent == "主调度泳道":
        return "可能旧式直投"
    missing = missing_loop_fields(msg)
    if missing:
        return "缺产品闭环字段"
    return "正常"


def clean_lane_label(value: Any, aliases: dict[str, str]) -> str:
    label = lane_label(value, aliases)
    if looks_garbled(label):
        raw = str(value or "")
        return LANE_LABELS.get(raw, "未知/乱码源")
    return label


LOOP_FIELDS = [
    "active_user_loop",
    "loop_impact",
    "user_loop_progress",
    "blocking_concerns",
    "backlog_concerns",
    "recommended_next_type",
]


def missing_loop_fields(msg: dict[str, Any]) -> list[str]:
    if msg.get("task_type") in TRANSPORT_TASK_TYPES:
        return []
    if not (completion_like(msg) or is_dispatch(msg)):
        return []
    if is_dispatch(msg):
        required = ["active_user_loop", "loop_impact"]
    elif str(msg.get("from_agent") or "") in {"review", "验收泳道"}:
        required = ["user_loop_progress", "blocking_concerns", "backlog_concerns", "recommended_next_type"]
    else:
        required = ["active_user_loop", "loop_impact", "blocking_concerns", "backlog_concerns", "recommended_next_type"]
    missing: list[str] = []
    for field in required:
        if field not in msg:
            missing.append(field)
            continue
        value = msg.get(field)
        if value is None or value == "":
            missing.append(field)
    return missing


def capability_key(msg: dict[str, Any]) -> str:
    raw = str(msg.get("slice") or msg.get("slice_id") or msg.get("task_type") or msg.get("message_id") or "")
    text = raw.lower()
    for prefix in ("value-slice-", "vs-"):
        text = text.replace(prefix, "")
    parts = [part for part in re.split(r"[^a-z0-9]+", text) if part]
    stop = {
        "001",
        "002",
        "review",
        "development",
        "design",
        "planning",
        "dispatch",
        "reply",
        "local",
        "consume",
        "consumer",
    }
    core = [part for part in parts if part not in stop]
    return "-".join(core[:4]) if core else raw[:80]


def recent_loop_warnings(messages: list[dict[str, Any]], window: int = 8) -> list[str]:
    work_items = [
        msg
        for msg in messages
        if (completion_like(msg) or is_dispatch(msg))
        and msg.get("task_type") not in TRANSPORT_TASK_TYPES
        and not message_is_garbled(msg)
    ][-window:]
    warnings: list[str] = []
    missing = [
        f"{msg.get('message_id')}: {', '.join(missing_loop_fields(msg))}"
        for msg in work_items
        if missing_loop_fields(msg)
    ]
    if missing:
        warnings.append("loop_fields_missing: " + " | ".join(missing[-5:]))

    counts: dict[str, int] = {}
    advanced_seen: dict[str, bool] = {}
    for msg in work_items:
        key = capability_key(msg)
        counts[key] = counts.get(key, 0) + 1
        advanced_seen[key] = advanced_seen.get(key, False) or msg.get("user_loop_progress") == "advanced" or str(
            msg.get("loop_impact") or ""
        ).startswith("advanced")
    for key, count in sorted(counts.items(), key=lambda item: item[1], reverse=True):
        if count >= 3 and not advanced_seen.get(key):
            warnings.append(
                f"possible_local_deep_dive: capability={key}, count_last_{window}={count}, no advanced user-loop signal"
            )
    return warnings


def communication_readable_row(index: int, msg: dict[str, Any], aliases: dict[str, str]) -> dict[str, str]:
    return {
        "序号": str(index),
        "时间": time_text(msg.get("created_at")),
        "类型": communication_type(msg),
        "来源": lane_label(msg.get("from_agent"), aliases),
        "去向": lane_label(msg.get("to_agent"), aliases),
        "状态": csv_value(msg.get("status")),
        "质量": quality_state(msg, aliases),
        "用户闭环": csv_value(msg.get("active_user_loop")),
        "闭环影响": csv_value(msg.get("loop_impact")),
        "用户闭环进展": csv_value(msg.get("user_loop_progress")),
        "建议下一步类型": csv_value(msg.get("recommended_next_type")),
        "闭环字段提醒": csv_value(missing_loop_fields(msg)),
        "对应任务": csv_value(msg.get("reply_to")),
        "消息编号": csv_value(msg.get("message_id")),
        "摘要": csv_value(msg.get("summary")),
        "建议动作": csv_value(msg.get("next_recommended_action")),
        "产物/文件": csv_value(msg.get("changed_files")),
        "验证证据": csv_value(msg.get("evidence")),
        "关注点": csv_value(msg.get("concerns")),
        "邮局批次": csv_value(msg.get("batch_id") or msg.get("outbox_path")),
        "主调度收到原文": csv_value(msg.get("thread_prompt") or msg.get("orchestrator_message")),
        "原始JSON": json.dumps(msg, ensure_ascii=False, separators=(",", ":")),
    }


def export_readable_xlsx(rows: list[dict[str, str]], path: Path) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except Exception:
        return False

    headers = list(rows[0].keys()) if rows else list(communication_readable_row(1, {}, {}).keys())
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "完整收发账本"
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    good_fill = PatternFill("solid", fgColor="E2F0D9")
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    bad_fill = PatternFill("solid", fgColor="FCE4D6")
    type_fill = PatternFill("solid", fgColor="DDEBF7")

    widths = {
        "序号": 8,
        "时间": 20,
        "类型": 16,
        "来源": 14,
        "去向": 14,
        "状态": 16,
        "质量": 18,
        "用户闭环": 46,
        "闭环影响": 22,
        "用户闭环进展": 22,
        "建议下一步类型": 26,
        "闭环字段提醒": 42,
        "对应任务": 36,
        "消息编号": 42,
        "摘要": 76,
        "建议动作": 76,
        "产物/文件": 52,
        "验证证据": 76,
        "关注点": 64,
        "邮局批次": 38,
        "主调度收到原文": 90,
        "原始JSON": 90,
    }

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    header_index = {name: idx + 1 for idx, name in enumerate(headers)}
    quality_idx = header_index.get("质量")
    type_idx = header_index.get("类型")
    for row in sheet.iter_rows(min_row=2):
        quality = str(row[quality_idx - 1].value or "") if quality_idx else ""
        row_fill = good_fill if quality == "正常" else warn_fill
        if "乱码" in quality or "重发" in quality:
            row_fill = bad_fill
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            cell.fill = row_fill
        if type_idx:
            row[type_idx - 1].fill = type_fill

    for column_cells in sheet.columns:
        header = str(column_cells[0].value or "")
        sheet.column_dimensions[column_cells[0].column_letter].width = widths.get(header, 20)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    for row_idx in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row_idx].height = 72
    sheet.row_dimensions[1].height = 28
    workbook.save(path)
    return True


def export_communications_workbook(messages: list[dict[str, Any]], aliases: dict[str, str]) -> Path:
    readable_xlsx_path = AGENT_LANES / "communications-readable.xlsx"
    rows = [communication_readable_row(index, msg, aliases) for index, msg in enumerate(messages, start=1)]
    export_readable_xlsx(rows, readable_xlsx_path)
    return readable_xlsx_path


def worklog_progress(path_value: Any) -> str:
    if not path_value:
        return "-"
    path = ROOT / str(path_value)
    if not path.exists():
        return "未找到 worklog"
    try:
        lines = path.read_text(encoding="utf-8-sig", errors="replace").splitlines()
    except UnicodeDecodeError:
        return "worklog 编码不可读"
    headings = [line.strip("# ").strip() for line in lines if line.startswith("## ")]
    if headings:
        return compact(headings[-1], 80)
    nonempty = [line.strip() for line in lines if line.strip()]
    return compact(nonempty[-1], 80) if nonempty else "暂无进度"


def msg_link(label: str, msg: dict[str, Any]) -> str:
    mid = compact(msg.get("message_id"), 72)
    return f"{label}: `{mid}`" if mid else label


def human_event(msg: dict[str, Any], aliases: dict[str, str]) -> str:
    task_type = str(msg.get("task_type") or "")
    source = lane_label(msg.get("from_agent"), aliases)
    status = str(msg.get("status") or "")
    summary = compact(msg.get("summary"), 140)
    if message_is_garbled(msg):
        return f"需要重发：检测到乱码回报 `{compact(msg.get('message_id'), 72)}`。"
    if task_type == "callback_batch_ready":
        count = len(msg.get("callback_ids") or [])
        return f"邮局把 {count} 条泳道回报合并成 1 条主调度消息。"
    if task_type == "callback_spooled":
        return f"{source} 的回报已进入邮局暂存，等待主调度空闲后合批。"
    if task_type == "orchestrator_state":
        state = "空闲" if msg.get("state") == "idle" else "忙碌"
        return f"主调度状态变为{state}：{compact(msg.get('reason'), 80)}"
    if completion_like(msg):
        return f"{source} 回报 {status}：{summary}"
    if is_dispatch(msg):
        return f"主调度派发：{summary or compact(task_type, 100)}"
    return summary or compact(task_type, 100)


def latest_post_office_status() -> dict[str, Any]:
    state_path = AGENT_LANES / "callback-inbox" / "post-office-state.json"
    pending_dir = AGENT_LANES / "callback-inbox" / "pending"
    state: dict[str, Any] = {}
    if state_path.exists():
        try:
            state = read_json(state_path)
        except json.JSONDecodeError:
            state = {"status": "unreadable"}
    state["pending_count"] = len(list(pending_dir.glob("*.json"))) if pending_dir.exists() else 0
    return state


def latest_outbox_status() -> dict[str, Any] | None:
    outbox_dir = AGENT_LANES / "callback-inbox" / "outbox"
    if not outbox_dir.exists():
        return None
    files = sorted(outbox_dir.glob("*-thread-send.json"), key=lambda path: path.stat().st_mtime)
    if not files:
        return None
    path = files[-1]
    try:
        item = read_json(path)
    except json.JSONDecodeError:
        item = {"status": "UNREADABLE"}
    item["outbox_path"] = path.relative_to(ROOT).as_posix()
    item.setdefault("delivery_state", "ready_to_send_or_retry")
    return item


def mark_historical_outbox(item: dict[str, Any] | None, current_state: dict[str, Any], pending_count: int) -> dict[str, Any] | None:
    if not item or pending_count:
        return item
    if str(item.get("status") or item.get("delivery_state") or "") not in {"READY_TO_SEND", "ready_to_send_or_retry"}:
        return item
    outbox_time = str(item.get("created_at") or "")
    state_time = str(current_state.get("updated_at") or "")
    if outbox_time and state_time and outbox_time < state_time:
        item["delivery_state"] = "historical_unacknowledged"
    return item


def first_lines(text: str, max_lines: int = 18) -> list[str]:
    lines = [line.rstrip() for line in text.splitlines() if line.strip()]
    if len(lines) <= max_lines:
        return lines
    return lines[:max_lines] + [f"...另有 {len(lines) - max_lines} 行，见批次消息文件。"]


def main() -> int:
    registry = read_json(AGENT_LANES / "agent-registry.json")
    legacy_rows = read_jsonl(AGENT_LANES / "message-log.jsonl")
    legacy_transport = [row for row in legacy_rows if row.get("task_type") in TRANSPORT_TASK_TYPES]
    messages = [row for row in legacy_rows if row.get("task_type") not in TRANSPORT_TASK_TYPES]
    transport_messages = legacy_transport + read_jsonl(AGENT_LANES / "transport-log.jsonl")
    all_events = messages + transport_messages
    current_state_path = AGENT_LANES / "current-state.json"
    current_state = read_json(current_state_path) if current_state_path.exists() else {}
    ledger_rows = read_jsonl(AGENT_LANES / "value-slice-ledger.jsonl")
    agents = registry.get("agents", [])

    lane_aliases: dict[str, str] = dict(LANE_LABELS)
    for agent in agents:
        agent_id = str(agent.get("agent_id", ""))
        display_name = str(agent.get("display_name", ""))
        if display_name and not message_is_garbled({"summary": display_name}):
            lane_aliases[agent_id] = display_name
            lane_aliases[display_name] = display_name

    known_lane_names = {lane_label(agent.get("agent_id"), lane_aliases) for agent in agents}
    callbacks = [msg for msg in messages if msg.get("task_type") in {"completion_callback", "completion_callback_protocol_update"} or completion_like(msg)]
    display_callbacks = [msg for msg in callbacks if valid_display_callback(msg)]
    dispatches = [msg for msg in messages if is_dispatch(msg)]
    user_requests_to_orchestrator = [msg for msg in dispatches if is_user_to_orchestrator_request(msg, lane_aliases)]
    lane_dispatches = [msg for msg in dispatches if not is_user_to_orchestrator_request(msg, lane_aliases)]
    current_active_dispatch = active_slice_dispatch(current_state)
    if current_active_dispatch and not any(
        msg.get("message_id") == current_active_dispatch.get("message_id") for msg in lane_dispatches
    ):
        lane_dispatches.append(current_active_dispatch)
    batches = [msg for msg in all_events if msg.get("task_type") == "callback_batch_ready"]
    orchestrator_states = [msg for msg in all_events if msg.get("task_type") == "orchestrator_state"]
    latest_batch = batches[-1] if batches else None
    latest_state = orchestrator_states[-1] if orchestrator_states else None
    post_office = latest_post_office_status()
    latest_outbox = mark_historical_outbox(
        latest_outbox_status(), current_state, int(post_office.get("pending_count", 0) or 0)
    )
    garbled_messages = [msg for msg in messages if message_is_garbled(msg)]
    current_garbled = [msg for msg in garbled_messages if str(msg.get("created_at") or "") >= "2026-06-24T15:49:29+08:00"]
    loop_warnings = recent_loop_warnings(messages)
    state_conflicts = control_plane_conflicts(current_state, ledger_rows, messages, lane_aliases)
    xlsx_path = export_communications_workbook(messages, lane_aliases)

    latest_callback_by_lane: dict[str, dict[str, Any]] = {}
    for msg in display_callbacks:
        latest_callback_by_lane[lane_label(msg.get("from_agent"), lane_aliases)] = msg

    latest_inbox_by_lane: dict[str, dict[str, Any]] = {}
    latest_open_dispatch_by_lane: dict[str, dict[str, Any]] = {}
    for msg in lane_dispatches:
        lane_name = lane_label(msg.get("to_agent"), lane_aliases)
        if lane_name not in known_lane_names:
            continue
        latest_inbox_by_lane[lane_name] = msg
        if not dispatch_is_covered_by_callback(messages, msg, lane_name, lane_aliases):
            latest_open_dispatch_by_lane[lane_name] = msg

    latest_spool_by_reply: dict[str, dict[str, Any]] = {}
    for msg in all_events:
        if msg.get("task_type") == "callback_spooled" and msg.get("reply_to"):
            latest_spool_by_reply[str(msg.get("reply_to"))] = msg

    actionable_open_dispatch_by_lane = {
        lane_name: msg
        for lane_name, msg in latest_open_dispatch_by_lane.items()
        if latest_inbox_by_lane.get(lane_name, {}).get("message_id") == msg.get("message_id")
    }
    if current_active_dispatch:
        active_dispatch_id = str(current_active_dispatch.get("message_id") or "")
        actionable_open_dispatch_by_lane = {
            lane_name: msg
            for lane_name, msg in actionable_open_dispatch_by_lane.items()
            if str(msg.get("message_id") or "") == active_dispatch_id
        }
    else:
        actionable_open_dispatch_by_lane = {}
    folded_historical_open_count = len(latest_open_dispatch_by_lane) - len(actionable_open_dispatch_by_lane)
    latest_user_request = user_requests_to_orchestrator[-1] if user_requests_to_orchestrator else None

    active_non_orchestrator = [
        agent for agent in agents if agent.get("agent_id") != "orchestrator" and agent.get("status") == "active"
    ]
    callback_covered = sum(
        1 for agent in active_non_orchestrator if lane_label(agent.get("agent_id"), lane_aliases) in latest_callback_by_lane
    )

    lines: list[str] = []
    lines.append("# Agent Lanes 总控台")
    lines.append("")
    lines.append(f"更新时间：`{now_text()}`")
    lines.append("")
    lines.append("这是给人看的入口：先看当前状态、邮局批次、泳道回报和下一步建议；长 ID、脚本路径和原始 JSON 放在最后的排查入口。")
    lines.append("")

    lines.append("## 一眼看懂")
    lines.append("")
    lines.append(f"- 活跃泳道：`{sum(1 for agent in agents if agent.get('status') == 'active')}`")
    lines.append(f"- 完成回报覆盖：`{callback_covered}/{len(active_non_orchestrator)}` 条非主调度活跃泳道已有回报记录")
    lines.append(
        f"- 泳道待回派发：`{len(actionable_open_dispatch_by_lane)}` 条当前最新收信未匹配到回报；"
        f"另有 `{folded_historical_open_count}` 条历史未折叠记录放入排查口径"
    )
    if latest_user_request:
        lines.append(
            f"- 主调度用户请求：最近 1 条记录为 `{compact(latest_user_request.get('message_id'), 60)}`；"
            "它不计入泳道待回，是否已处理以主调度 worklog / 后续派发为准"
        )
    if current_garbled:
        lines.append(f"- 当前质量警告：`{len(current_garbled)}` 条新记录含乱码，需对应泳道重发 UTF-8 callback")
    else:
        lines.append("- 当前质量警告：无新乱码回报")
    if loop_warnings:
        lines.append(f"- 产品闭环节奏警告：`{len(loop_warnings)}` 条，见下方 `Product Loop 节奏警告`")
    else:
        lines.append("- 产品闭环节奏：正常，未发现同能力深挖或关键字段缺失")
    if state_conflicts:
        lines.append(f"- 控制面冲突：`{len(state_conflicts)}` 条；自动派发必须停止，详见下方")
    else:
        lines.append("- 控制面一致性：正常")
    if latest_state:
        state_label = "空闲" if latest_state.get("state") == "idle" else "忙碌"
        lines.append(f"- 主调度状态：{state_label}（{compact(latest_state.get('reason'), 80)}）")
    else:
        lines.append("- 主调度状态：未看到 `orchestrator_state`，邮局会谨慎等待")
    po_action = compact(post_office.get("last_action") or post_office.get("status") or "unknown", 50)
    lines.append(f"- 回报邮局：{po_action}，暂存待投递 `pending={post_office.get('pending_count', 0)}`")
    if latest_outbox:
        outbox_status = str(latest_outbox.get("delivery_state") or latest_outbox.get("status") or "READY_TO_SEND")
        if outbox_status in {"READY_TO_SEND", "ready_to_send_or_retry"}:
            outbox_label = "已生成待发送 / 待重试"
        elif outbox_status == "historical_unacknowledged":
            outbox_label = "历史未确认投递（不触发业务重做）"
        elif outbox_status in {"SENT", "DELIVERED", "HANDLED"}:
            outbox_label = "已由主调度处理或已发送"
        else:
            outbox_label = outbox_status
        message_path = latest_outbox.get("orchestrator_message_path") or latest_outbox.get("outbox_path")
        lines.append(
            f"- 邮局投递状态：{outbox_label}；批次 `{compact(latest_outbox.get('batch_id'), 44)}`；"
            f"目标线程 `{compact(latest_outbox.get('target_thread_id'), 36)}`；"
            f"{md_link('合并信', message_path)} / {md_link('outbox', latest_outbox.get('outbox_path', ''))}"
        )
    if latest_batch:
        lines.append(
            f"- 最近 message-log 邮局批次：{len(latest_batch.get('callback_ids') or [])} 条回报已合并为 1 条主调度消息；"
            f"`{compact(latest_batch.get('message_id'), 48)}`"
        )
    lines.append(f"- 完整收发账本：{md_link('推荐打开 Excel 版', xlsx_path)}")
    lines.append("")

    lines.append("## 泳道收发工作台")
    lines.append("")
    lines.append("| 泳道 | 最近收信 | 当前进度 | 最近回信 | 邮局/质量 |")
    lines.append("| --- | --- | --- | --- | --- |")
    for agent in agents:
        agent_id = str(agent.get("agent_id") or "")
        display = lane_label(agent_id, lane_aliases)
        inbox = latest_inbox_by_lane.get(display)
        open_dispatch = actionable_open_dispatch_by_lane.get(display)
        callback = latest_callback_by_lane.get(display)
        inbox_text = "-"
        if inbox:
            state = "可能待回" if open_dispatch and open_dispatch.get("message_id") == inbox.get("message_id") else "已处理/有后续"
            inbox_text = f"{state}<br>{msg_link('收信', inbox)}<br>{compact(inbox.get('summary'), 92)}"
        progress = worklog_progress(agent.get("worklog"))
        reply_text = "-"
        quality_bits: list[str] = []
        if callback:
            reply_text = f"`{short_time(callback.get('created_at'))}` {callback.get('status', '-')}<br>{msg_link('回信', callback)}<br>{compact(callback.get('summary'), 92)}"
            spool = latest_spool_by_reply.get(str(callback.get("message_id") or ""))
            if callback.get("batch_id"):
                quality_bits.append(f"邮局批次 `{compact(callback.get('batch_id'), 34)}`")
            elif spool:
                quality_bits.append("已进邮局暂存")
            elif display != "主调度泳道":
                quality_bits.append("未见邮局批次")
        lane_messages = [
            msg
            for msg in messages
            if lane_label(msg.get("from_agent"), lane_aliases) == display or lane_label(msg.get("to_agent"), lane_aliases) == display
        ]
        lane_current_garbled = [msg for msg in lane_messages if msg in current_garbled]
        if lane_current_garbled:
            quality_bits.append(f"乱码警告 {len(lane_current_garbled)}")
        if not quality_bits:
            quality_bits.append("正常")
        lines.append(f"| {display} | {inbox_text} | {progress} | {reply_text} | {'<br>'.join(quality_bits)} |")
    lines.append("")

    if current_garbled:
        lines.append("## 当前质量警告")
        lines.append("")
        lines.append("这些记录是在乱码防线启用后出现的，不能当作有效中文回报。主视图已隔离原文；如需真实语义，应让对应泳道重新生成 UTF-8 callback。")
        lines.append("")
        lines.append("| 时间 | 记录 | 来源 | 处理方式 |")
        lines.append("| --- | --- | --- | --- |")
        for msg in current_garbled[-12:][::-1]:
            source = clean_lane_label(msg.get("from_agent"), lane_aliases)
            lines.append(
                f"| `{short_time(msg.get('created_at'))}` | `{compact(msg.get('message_id'), 72)}` | {source} | 已隔离乱码原文，保留审计；需要时重发 UTF-8 callback |"
            )
        lines.append("")

    if loop_warnings:
        lines.append("## Product Loop 节奏警告")
        lines.append("")
        lines.append("这些警告用于防止系统重新滑回局部深挖：字段缺失要补齐；同一 capability 连续出现但没有 `advanced` 信号时，主调度应先做 Product Loop Check。")
        lines.append("")
        for warning in loop_warnings:
            lines.append(f"- `{warning}`")
        lines.append("")

    if latest_batch:
        lines.append("## 最近一封邮局合并信")
        lines.append("")
        message_text = str(latest_batch.get("orchestrator_message") or "")
        if message_text and not looks_garbled(message_text):
            for line in first_lines(message_text, 18):
                lines.append(f"> {line}")
        elif message_text:
            lines.append("> 最近批次原文含乱码，dashboard 已隔离正文，避免污染人看的主视图。")
            lines.append("> 请以 `message_id` 追溯审计，或让对应泳道重发 UTF-8 callback。")
        else:
            count = len(latest_batch.get("callback_ids") or [])
            lines.append(f"> 本批共 {count} 条回报，旧记录未包含合并正文。")
        if latest_batch.get("orchestrator_message_path"):
            lines.append("")
            lines.append(f"完整批次消息：{md_link('打开合并信', latest_batch['orchestrator_message_path'])}")
        lines.append("")

    lines.append("## 泳道状态")
    lines.append("")
    lines.append("| 泳道 | 状态 | 最近回报 | 下一步建议 | 详情 |")
    lines.append("| --- | --- | --- | --- | --- |")
    for agent in agents:
        agent_id = str(agent.get("agent_id") or "")
        display = lane_label(agent_id, lane_aliases)
        callback = latest_callback_by_lane.get(display)
        callback_text = "-"
        next_action = "-"
        if callback:
            callback_text = f"{callback.get('status', '-')}: {compact(callback.get('summary'), 76)}"
            next_action = compact(callback.get("next_recommended_action"), 76) or lane_label(
                callback.get("next_recommended_lane"), lane_aliases
            )
        detail = f"{md_link('worklog', agent.get('worklog', ''))} / {md_link('workspace', agent.get('workspace', ''))}"
        lines.append(f"| {display} | {status_badge(agent.get('status'))} | {callback_text} | {next_action} | {detail} |")
    lines.append("")

    lines.append("## 近期有效产物记录")
    lines.append("")
    lines.append("这里展示给人读的完整记录：只放未乱码的有效 completion callback；历史乱码记录放到质量/排查区，不在这里当作正常产物。")
    lines.append("")
    if display_callbacks:
        for msg in display_callbacks[-6:][::-1]:
            source = lane_label(msg.get("from_agent"), lane_aliases)
            lines.append(f"### `{short_time(msg.get('created_at'))}` {source} · `{msg.get('status', '-')}`")
            lines.append("")
            lines.append(f"- 回报编号：`{compact(msg.get('message_id'), 96)}`")
            if msg.get("reply_to"):
                lines.append(f"- 对应任务：`{compact(msg.get('reply_to'), 96)}`")
            lines.append("")
            append_value_block(lines, "完成摘要", msg.get("summary"))
            append_value_block(lines, "产物 / 变更文件", msg.get("changed_files"))
            append_value_block(lines, "验证证据", msg.get("evidence"))
            append_value_block(lines, "关注点", msg.get("concerns"))
            next_action = msg.get("next_recommended_action") or lane_label(msg.get("next_recommended_lane"), lane_aliases)
            append_value_block(lines, "建议动作", next_action)
    else:
        lines.append("暂无有效完成回报。")
        lines.append("")

    historical_garbled_callbacks = [msg for msg in callbacks if message_is_garbled(msg)]
    if historical_garbled_callbacks:
        lines.append("## 历史质量污染记录")
        lines.append("")
        lines.append("这些记录保留为审计事实，但不再作为正常产物展示；如需真实语义，应让对应泳道重发 UTF-8 callback。")
        lines.append("")
        lines.append("| 时间 | 记录 | 来源 | 处理方式 |")
        lines.append("| --- | --- | --- | --- |")
        for msg in historical_garbled_callbacks[-8:][::-1]:
            source = clean_lane_label(msg.get("from_agent"), lane_aliases)
            lines.append(f"| `{short_time(msg.get('created_at'))}` | `{compact(msg.get('message_id'), 84)}` | {source} | 已从主产物区过滤，仅保留审计 |")
        lines.append("")

    lines.append("## 最近动态")
    lines.append("")
    lines.append("| 时间 | 动态 |")
    lines.append("| --- | --- |")
    for msg in messages[-15:][::-1]:
        lines.append(f"| `{short_time(msg.get('created_at'))}` | {human_event(msg, lane_aliases)} |")
    if not messages:
        lines.append("| - | 暂无消息 |")
    lines.append("")

    smoke_paths = [
        ("规划 smoke 产物", "agent-lanes/lanes/planning/workspace/smoke-test-plan.md"),
        ("设计 smoke 产物", "agent-lanes/lanes/design/workspace/smoke-test-empty-state-design.md"),
        ("验收 smoke 结论", "agent-lanes/lanes/review/workspace/smoke-test-review.md"),
    ]
    existing_smoke = [(label, path) for label, path in smoke_paths if (ROOT / path).exists()]
    if existing_smoke:
        lines.append("## 最近闭环产物")
        lines.append("")
        for label, path in existing_smoke:
            lines.append(f"- {md_link(label, path)}")
        lines.append("")

    lines.append("## 排查入口")
    lines.append("")
    lines.append("- 人看状态：本文件。")
    lines.append("- 机器追溯：`agent-lanes/message-log.jsonl`。")
    lines.append(f"- 推荐阅读账本：{md_link('agent-lanes/communications-readable.xlsx', xlsx_path)}。")
    lines.append("- 邮局暂存：`agent-lanes/callback-inbox/pending/`。")
    lines.append("- 邮局已投递批次：`agent-lanes/callback-inbox/delivered/`。")
    lines.append("- 刷新命令：`python agent-lanes/scripts/render_dashboard.py`。")
    lines.append("")

    full_dashboard = AGENT_LANES / "dashboard-full.md"
    full_dashboard.write_text("\n".join(lines) + "\n", encoding="utf-8-sig", newline="\n")

    compact_lines: list[str] = []
    compact_lines.append("# Agent Lanes 当前状态")
    compact_lines.append("")
    compact_lines.append(f"更新时间：`{now_text()}`")
    compact_lines.append("")
    compact_lines.append("本文件只展示当前决策状态；完整审计与历史回报见 `dashboard-full.md`、`message-log.jsonl` 和 `transport-log.jsonl`。")
    compact_lines.append("")
    compact_lines.append("## 当前价值切片")
    compact_lines.append("")
    compact_lines.append(f"- 北极星：{current_state.get('north_star', '未登记')}")
    compact_lines.append(f"- 机制版本：`{current_state.get('mechanism_version', '-')}` / `{current_state.get('mechanism_status', '-')}`")
    active_slice = current_state.get("active_value_slice")
    if isinstance(active_slice, dict):
        compact_lines.append(f"- Value Slice：`{active_slice.get('value_slice_id', '-')}` · {active_slice.get('title', '-')}")
        compact_lines.append(f"- 用户价值变化：{active_slice.get('new_user_action', '-')}")
    else:
        compact_lines.append("- Value Slice：当前没有已通过 Product Value Gate 的业务派发")
    compact_lines.append(f"- 当前用户闭环：{current_state.get('active_user_loop', '-')}")
    compact_lines.append(f"- 建议下一价值切片：{current_state.get('next_recommended_value_slice', '-')}")
    automation_policy = current_state.get("automation_policy") or {}
    if automation_policy:
        compact_lines.append(
            f"- 受控自动推进：`{automation_policy.get('status', '-')}` · "
            f"连续切片 `{automation_policy.get('consecutive_auto_slices', 0)}/{automation_policy.get('max_consecutive_auto_slices', 0)}` · "
            f"迭代 `{automation_policy.get('auto_iteration_id', '-')}`"
        )
    candidates = current_state.get("candidate_slices") or []
    for candidate in candidates[:2]:
        compact_lines.append(
            f"- 候选 `{candidate.get('rank', '-')}`：`{candidate.get('value_slice_id', '-')}` · {candidate.get('title', '-')}"
        )
    compact_lines.append("")
    compact_lines.append("## 调度状态")
    compact_lines.append("")
    compact_lines.append(f"- 已注册泳道：`{len(agents)}`；当前待回派发：`{len(actionable_open_dispatch_by_lane)}`")
    compact_lines.append(f"- 业务事件：`{len(messages)}`；传输事件：`{len(transport_messages)}`")
    compact_lines.append(f"- Product Loop 警告：`{len(loop_warnings)}`；乱码警告：`{len(current_garbled)}`")
    compact_lines.append(f"- 控制面冲突：`{len(state_conflicts)}`；非零时自动派发必须停止")
    if latest_state:
        compact_lines.append(f"- 主调度：`{latest_state.get('state', '-')}` · {compact(latest_state.get('reason'), 100)}")
    else:
        compact_lines.append("- 主调度：未登记新的 V2 transport state")
    compact_lines.append(f"- 邮局：`{po_action}`；pending=`{post_office.get('pending_count', 0)}`")
    if latest_outbox:
        compact_lines.append(
            f"- 最新 outbox：`{latest_outbox.get('delivery_state', latest_outbox.get('status', '-'))}` · "
            f"{md_link('打开', latest_outbox.get('outbox_path', ''))}"
        )
    compact_lines.append("")
    compact_lines.append("## 当前阻塞与 backlog")
    compact_lines.append("")
    blockers = current_state.get("open_blocking_concerns") or []
    if blockers:
        compact_lines.extend(f"- BLOCKING：{item}" for item in blockers)
    else:
        compact_lines.append("- BLOCKING：无")
    for item in state_conflicts:
        compact_lines.append(f"- BLOCKING（控制面冲突）：{item}")
    for item in (current_state.get("backlog_concerns") or [])[:5]:
        compact_lines.append(f"- BACKLOG：{item}")
    compact_lines.append("")
    compact_lines.append("## 泳道当前任务")
    compact_lines.append("")
    compact_lines.append("| 泳道 | 当前待回 | 最近有效回报 |")
    compact_lines.append("| --- | --- | --- |")
    for agent in agents:
        display = lane_label(agent.get("agent_id"), lane_aliases)
        open_dispatch = actionable_open_dispatch_by_lane.get(display)
        callback = latest_callback_by_lane.get(display)
        pending_text = compact(open_dispatch.get("summary") or open_dispatch.get("task_type"), 68) if open_dispatch else "无"
        callback_text = f"{callback.get('status')}: {compact(callback.get('summary'), 68)}" if callback else "无"
        compact_lines.append(f"| {display} | {pending_text} | {callback_text} |")
    compact_lines.append("")
    compact_lines.append("## 最近业务结果")
    compact_lines.append("")
    for msg in display_callbacks[-4:][::-1]:
        compact_lines.append(
            f"- `{short_time(msg.get('created_at'))}` {lane_label(msg.get('from_agent'), lane_aliases)} · "
            f"`{msg.get('status', '-')}` · {compact(msg.get('summary'), 120)}"
        )
    if not display_callbacks:
        compact_lines.append("- 暂无有效业务回报")
    compact_lines.append("")
    compact_lines.append("## 派发门禁")
    compact_lines.append("")
    compact_lines.append("- 新任务必须绑定 `agent-lanes.value-slice.v2`。")
    compact_lines.append("- 正式派发：`python agent-lanes/scripts/product_value_gate.py --dispatch-file <dispatch.json> --record-dispatch`。")
    compact_lines.append("- 只做候选验证：显式使用 `--dry-run`；不得把 dry-run 结果当成已派发。")
    compact_lines.append("- `recommended_next_*` 只作咨询；门禁通过后才允许派发。")
    compact_lines.append("- 完整状态：`agent-lanes/dashboard-full.md`；业务审计：`message-log.jsonl`；传输排查：`transport-log.jsonl`。")
    compact_lines.append("")

    dashboard = AGENT_LANES / "dashboard.md"
    # UTF-8 with BOM keeps Windows/desktop Markdown previews from misreading Chinese as ANSI.
    dashboard.write_text("\n".join(compact_lines) + "\n", encoding="utf-8-sig", newline="\n")
    print(f"Rendered {dashboard} and {full_dashboard}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
